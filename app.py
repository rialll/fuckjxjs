from flask import Flask, request
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

file_path = r'C:\Users\Rialll\Desktop\tk.xlsx'  # 文件路径
sheet_name = 'Sheet2 '  # 工作表名称，
b_column_index = 1  # 题目列 B 列的索引为 1（从 0 开始）
h_column_index = 7  # 答案列 H 列的索引为 7（从 0 开始）

def find_h_value(search_text):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows():
        cell_value = row[b_column_index].value
        if cell_value == search_text:
            h_value = row[h_column_index].value
            return str(h_value)

    return ''

@app.route('/', methods=['POST'])
def get_h_value():
    search_text = request.form.get('title')
    h_value = find_h_value(search_text)
    return h_value

if __name__ == '__main__':
    app.run()
